# -*- coding: utf-8 -*-
# -----------------------------------------------------------------------------
# Name:         urls.py (App Level)
# Purpose:      URL routing for the Invoice application's web pages.
#
# Author:       AnikRoy
# GitHub:       https://github.com/aroyslipk
#
# Created:      2025-06-20
# Copyright:    (c) AnikRoy 2025
# Licence:      Proprietary
# -----------------------------------------------------------------------------


from decimal import Decimal
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Protection
from django.conf import settings
import os
from num2words import num2words
from .models import WorkEntry, ClientProject

def generate_bank_invoice(request, project_id=None):
    project_id = request.GET.get('project')
    if not project_id:
        return HttpResponse("Project ID is required", status=400)

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Get work entries for the project
    entries = WorkEntry.objects.filter(project_id=project_id)
    if start_date:
        entries = entries.filter(date__gte=start_date)
    if end_date:
        entries = entries.filter(date__lte=end_date)

    # Calculate totals
    total_units = sum(entry.quantity for entry in entries)
    total_amount = sum(entry.quantity * entry.category.rate if entry.category else 0 for entry in entries)

    # Load template
    template_path = os.path.join(settings.STATICFILES_DIRS[0], 'template', 'InvoiceTemplate.xlsx')
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Format date range for invoice number
    date_range = ""
    if start_date and end_date:
        date_range = f"{start_date} to {end_date}"
    elif start_date:
        date_range = f"From {start_date}"
    elif end_date:
        date_range = f"Until {end_date}"

    # Set invoice details
    ws['B8'] = "Date:"  # Date label
    ws['C8'] = entries.first().date.strftime('%Y-%m-%d') if entries.exists() else ''  # Date
    ws['B9'] = "Invoice No:"  # Invoice number label
    ws['C9'] = f"Bank Invoice - {date_range}"  # Invoice number

    # Write totals
    ws['B13'] = "Total Units:"
    ws['C13'] = total_units
    ws['D13'] = "Total Amount (USD):"
    ws['E13'] = total_amount
    
    # Convert amount to words
    amount_words = num2words(total_amount, lang='en', to='currency', currency='USD').title()
    ws['B15'] = f"In Words: {amount_words}"
    ws.merge_cells('B15:E15')
    
    # Create bordered box for bank details
    border = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )
    
    # Set bank details with box and bold text
    bold_font = Font(bold=True, size=10, name='Arial')
    ws['B19'] = "BANK DETAILS"
    ws['B19'].font = Font(bold=True, size=12, name='Arial')
    
    bank_details = [
        "Account Name: Offshore Clipping",
        "Account Number: XXXX-XXXX-XXXX",
        "Bank Name: Your Bank",
        "SWIFT Code: XXXXXX"
    ]
    
    # Apply bank details with formatting
    for i, detail in enumerate(bank_details, start=20):
        ws[f'B{i}'] = detail
        cell = ws[f'B{i}']
        cell.font = bold_font
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Make sure all cells are unlocked for editing
    ws.protection.sheet = False
    unlocked = Protection(locked=False)
    for row in ws.rows:
        for cell in row:
            cell.protection = unlocked

    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=bank_invoice.xlsx'
    wb.save(response)
    return response
